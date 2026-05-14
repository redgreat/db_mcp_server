from typing import Dict, Any, List, Optional, Tuple
from sqlalchemy.engine import Engine
from sqlalchemy import text
from io import BytesIO, StringIO
import csv
import time
import os

def _safe_identifier(name: str) -> str:
    return "".join(ch for ch in name if ch.isalnum() or ch in ("_", ".", " ")).strip()

def _build_select_sql(source: Dict[str, Any], fields: List[Dict[str, Any]]) -> str:
    select_parts = []
    for f in fields:
        expr = f.get("expr") or f.get("column") or ""
        alias = f.get("alias") or f.get("label") or None
        expr = expr.strip()
        if not expr:
            continue
        if alias:
            select_parts.append(f"{expr} AS {alias}")
        else:
            select_parts.append(expr)
    if not select_parts:
        raise ValueError("empty fields")
    base_table = source.get("table")
    if not base_table:
        raise ValueError("missing table")
    sql = f"SELECT {', '.join(select_parts)} FROM {base_table}"
    joins = source.get("joins") or []
    for j in joins:
        jt = (j.get("type") or "LEFT").upper()
        jt = jt if jt in ("LEFT", "RIGHT", "INNER", "FULL", "LEFT OUTER", "RIGHT OUTER") else "LEFT"
        tbl = j.get("table")
        on = j.get("on")
        if tbl and on:
            sql += f" {jt} JOIN {tbl} ON {on}"
    where = source.get("where")
    if where:
        sql += f" WHERE {where}"
    order_by = source.get("order_by")
    if order_by:
        sql += f" ORDER BY {order_by}"
    limit = source.get("limit")
    if isinstance(limit, int) and limit > 0:
        sql += f" LIMIT {limit}"
    return sql

def _rows_to_csv_bytes(rows: List[Dict[str, Any]], headers: List[str]) -> bytes:
    sio = StringIO()
    writer = csv.writer(sio)
    writer.writerow(headers)
    for r in rows:
        writer.writerow([r.get(h) for h in headers])
    data = sio.getvalue().encode("utf-8")
    return data

def _rows_to_xlsx_bytes(rows: List[Dict[str, Any]], headers: List[str], field_defs: List[Dict[str, Any]]) -> bytes:
    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
    except Exception:
        return _rows_to_csv_bytes(rows, headers)
    wb = Workbook()
    ws = wb.active
    ws.append(headers)
    for r in rows:
        ws.append([r.get(h) for h in headers])
    for idx, f in enumerate(field_defs, start=1):
        w = f.get("width")
        if isinstance(w, int) and w > 0:
            ws.column_dimensions[get_column_letter(idx)].width = w
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()

def generate_and_upload_report(
    eng: Engine,
    access_key: str,
    cfg: Any,
    args: Dict[str, Any],
) -> Dict[str, Any]:
    from ..security.interceptor import intercept_sql
    from .db_tool import run_select
    from .object_storage import ObjectStorageClient
    source = args.get("source") or {}
    fields = args.get("fields") or []
    if not fields:
        raise ValueError("fields required")
    sql = _build_select_sql(source, fields)
    sec = intercept_sql(sql, {"key": access_key})
    if not sec.get("safe"):
        raise RuntimeError(f"risk sql: {sec.get('risk')}")
    rows = run_select(eng, sql)
    headers = []
    for f in fields:
        label = f.get("label") or f.get("alias") or f.get("expr") or f.get("column")
        headers.append(label)
    fmt = (args.get("output_format") or "xlsx").lower()
    ts = time.strftime("%Y%m%d_%H%M%S")
    name = args.get("file_name") or f"report_{ts}"
    if fmt not in ("xlsx", "csv"):
        fmt = "xlsx"
    if fmt == "xlsx":
        content = _rows_to_xlsx_bytes(rows, headers, fields)
        ext = "xlsx" if content[:4] != b"sep=" else "csv"
        if ext == "xlsx":
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        else:
            content_type = "text/csv; charset=utf-8"
    else:
        content = _rows_to_csv_bytes(rows, headers)
        ext = "csv"
        content_type = "text/csv; charset=utf-8"
    osc = None
    if cfg.object_storage and cfg.object_storage.enabled:
        osc = ObjectStorageClient(
            provider=cfg.object_storage.provider,
            endpoint=cfg.object_storage.endpoint,
            bucket=cfg.object_storage.bucket,
            access_key_id=cfg.object_storage.access_key_id,
            access_key_secret=cfg.object_storage.access_key_secret,
            region=cfg.object_storage.region,
            public_base_url=cfg.object_storage.public_base_url,
            base_path=cfg.object_storage.path_prefix,
        )
    else:
        osc = ObjectStorageClient(provider="local", endpoint=os.path.join("uploads"))
    obj_key = f"{name}.{ext}"
    url, final_key = osc.upload_bytes(obj_key, content, content_type=content_type)
    return {
        "url": url,
        "object_key": final_key,
        "format": "xlsx" if content_type.startswith("application/vnd") else "csv",
        "row_count": len(rows),
        "headers": headers,
        "sql": sql,
    }
